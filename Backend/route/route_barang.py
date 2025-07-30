from fastapi import APIRouter, Depends, HTTPException,Query,File,Form,UploadFile
from sqlmodel import select , Session 
from database import get_session, engine 
from model.model_barang import Barang, SQLModel ,BarangUpdate , BarangCreate , Histori , HistoriCreate , HistoriRead
from auth import AuthHandler
from fastapi.responses import FileResponse ,StreamingResponse
from openpyxl import Workbook
import os
from openpyxl.styles import Font, Border, Side, Alignment
import qrcode
from sqlalchemy import func
from io import BytesIO
from datetime import datetime
from typing import List , Optional
from uuid import uuid4
import shutil
from PIL import Image   
from dateutil.relativedelta import relativedelta
from datetime import date


auth_handler = AuthHandler()
barang = APIRouter()


@barang.post("/addhistori", response_model=HistoriRead)
def add_histori(histori_data: HistoriCreate, session: Session = Depends(get_session)):
    histori = Histori(
        nama=histori_data.nama,
        terjual=histori_data.terjual,
        harga=histori_data.harga,
        waktujual=histori_data.waktujual or datetime.utcnow()
    )
    session.add(histori)
    session.commit()
    session.refresh(histori)
    return histori

@barang.get("/histori", response_model=list[HistoriRead])
def get_histori(session: Session = Depends(get_session)):
    return session.exec(select(Histori)).all()




@barang.get("/{nama}/qrcode")
def generate_qrcode(nama: str, session: Session = Depends(get_session)):
    
    print("Mencari barang dengan nama:", nama)
    
    barang_data = session.exec(
        select(Barang).where(func.lower(Barang.nama) == nama.lower())
    ).first()
    
    print("Hasil pencarian:", barang_data)

    if not barang_data:
        return {"error": "Barang tidak ditemukan"}

    qr_data = f"Nama Barang: {barang_data.nama}\nHarga Barang: Rp{barang_data.harga}\nLokasi Penempatan Barang: {barang_data.lokasi}\nWaktu Pembelian Barang: {barang_data.waktu}\nJenis Barang: {barang_data.jenis}"

    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=4,
    )
    qr.add_data(qr_data)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")

    img_io = BytesIO()
    img.save(img_io, format='PNG')
    img_io.seek(0)

    return StreamingResponse(img_io, media_type="image/png")


@barang.get("/laporan/export-excel")
def export_barang_excel(session: Session = Depends(get_session)):
    barangs = session.exec(select(Barang)).all()
    
    waktu = datetime.now()
    tgl= waktu.date()
    wb = Workbook()
    ws = wb.active
    ws.title = f"Laporan Barang Tanggal {tgl}"
    

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ["Nama Barang", "Harga Barang(Rp)", "Stok Barang(Pcs)", "Lokasi Penempatan Barang" , "Waktu Pembelian Barang" ,"Usia Barang" , "jenis Barang" , "Barang Terjual (Pcs)"]
    ws.append(headers)

    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for row_num, item in enumerate(barangs, 2):
        sekarang = datetime.now().date()
        selisih = relativedelta(sekarang, item.waktu)

        hari = selisih.days
        bulan = selisih.months
        tahun = selisih.years

        usia = f"{hari} Hari, {bulan} Bulan, {tahun} Tahun"

            
        row_data = [item.nama, item.harga, item.stok, item.lokasi , item.waktu , usia ,  item.jenis , item.terjual]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='left')

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=laporan_barang.xlsx"}
    )

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploaded_images")
os.makedirs(UPLOAD_DIR, exist_ok=True)

@barang.post("/tambah")
async def tambah_barang(
    nama: str = Form(...),
    harga: int = Form(...),
    stok: int = Form(...),
    lokasi: str = Form(...),
    waktu: str = Form(...),
    jenis: Optional[str] = Form(None),
    gambar: UploadFile = File(None),
    session: Session = Depends(get_session)
):
    if not nama or harga < 0 or stok < 0 or not lokasi or not waktu:
        raise HTTPException(status_code=400, detail="Data tidak lengkap atau tidak valid")

    try:
        waktu_date = datetime.strptime(waktu, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Format tanggal tidak valid, gunakan YYYY-MM-DD")

    filename = None
    if gambar:
        ext = gambar.filename.split(".")[-1].lower()
        if ext not in {"jpg", "jpeg", "png"}:
            raise HTTPException(status_code=400, detail="Format gambar harus jpg/jpeg/png")

        filename = f"{uuid4()}.{ext}"
        filepath = os.path.join(UPLOAD_DIR, filename)

        try:
            with open(filepath, "wb") as f:
                gambar.file.seek(0)
                shutil.copyfileobj(gambar.file, f)

            with Image.open(filepath) as img:
                img.thumbnail((800, 800)) 
                img.save(filepath)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Gagal menyimpan atau resize gambar: {str(e)}")

    barang_baru = Barang(
        nama=nama,
        harga=harga,
        stok=stok,
        lokasi=lokasi,
        waktu=waktu_date,
        jenis=jenis,
        gambar=filename
    )
    session.add(barang_baru)
    session.commit()
    session.refresh(barang_baru)
    return barang_baru


@barang.get("/ambil", response_model=List[Barang])
def get_barang(
    skip: int = Query(0, ge=0),
    limit: int = Query(10, ge=1),
    jenis: Optional[str] = None,
    session: Session = Depends(get_session)
):
    query = select(Barang)
    if jenis:
        query = query.where(Barang.jenis == jenis)
    barang = session.exec(query.offset(skip).limit(limit)).all()
    return barang

@barang.put("/{id}/jual")
def jual_barang(id: int):
    with Session(engine) as session:
        barang = session.get(Barang, id)
        if not barang:
            raise HTTPException(status_code=404, detail="Barang tidak ditemukan")

        if barang.stok <= 0:
            raise HTTPException(status_code=400, detail="Stok barang habis")
        
        barang.waktujual = date.today()
        barang.stok -= 1
        barang.terjual += 1 
        session.add(barang)
        session.commit()
        session.refresh(barang)
        return barang


@barang.put("/update/{nama}")
def update_barang(nama: str, updated_barang: BarangUpdate,session = Depends(get_session)):
    existing_barang = session.exec(select(Barang).where(Barang.nama == nama)).first()

    if not existing_barang:
        raise HTTPException(status_code=404, detail="Barang tidak ditemukan")

    if updated_barang.harga is not None:
        existing_barang.harga = updated_barang.harga
    if updated_barang.stok is not None:
        existing_barang.stok = updated_barang.stok

    session.add(existing_barang)
    session.commit()
    session.refresh(existing_barang)

    return {"detail": f"Barang '{nama}' berhasil diperbarui"}


@barang.delete("/hapus/{nama}")
def delete_barang(nama:str, session=Depends(get_session)):
    existing_barang = session.exec(select(Barang).where(Barang.nama == nama)).first()
    if not existing_barang:
        raise HTTPException(status_code=404, detail="Barang tak ditemukan")
    
    session.delete(existing_barang)         
    session.commit()
    return {"detail": f"{existing_barang.nama} berhasil dihapus"}
